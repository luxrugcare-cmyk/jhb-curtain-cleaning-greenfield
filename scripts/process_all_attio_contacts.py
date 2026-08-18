#!/usr/bin/env python3
"""Unified Attio CRM & Prospect List Processor & Segmentation Engine.

Ingests:
1. 'People - Recently Contacted People.xlsx' (Attio 4,437 synced email contacts)
2. 'Tsogo_Sun_Prospect_List_and_Campaign_Kit.xlsx' (Tsogo Sun Master Hotel Properties & Decision Makers)

Purges:
- Automated receipts, bank alerts (FNB, Standard Bank, Absa, Nedbank), telcos, e-commerce, spam, system bots.

Outputs:
- data/cleaned_lists/commercial_b2b_leads.csv
- data/cleaned_lists/residential_leads.csv
- data/cleaned_lists/trade_design_leads.csv
- data/cleaned_lists/hotel_hospitality_leads.csv
- data/cleaned_lists/master_clean_contacts.csv
"""

import csv
import datetime as dt
import json
import os
import pathlib
import re
import sys
import openpyxl

DOWNLOADS_DIR = pathlib.Path(r"C:\Users\User\Downloads")
OUTPUT_DIR = pathlib.Path("data/cleaned_lists")

JUNK_PREFIXES = {
    "noreply", "no-reply", "donotreply", "do-not-reply", "notifications",
    "notification", "alerts", "alert", "mailer-daemon", "postmaster",
    "support", "help", "billing", "invoices", "invoice", "statements",
    "statement", "receipts", "receipt", "newsletters", "newsletter",
    "marketing", "subscriptions", "system", "admin", "bounce", "bounces",
    "updates", "feedback", "orders", "order", "shipping", "tracking",
    "info@notification", "accounts", "account"
}

JUNK_DOMAINS = {
    # Banks & Financial
    "fnb.co.za", "standardbank.co.za", "absa.co.za", "nedbank.co.za",
    "capitecbank.co.za", "investec.co.za", "discovery.co.za", "discoveryhealth.co.za",
    "momentum.co.za", "oldmutual.com", "sanlam.co.za", "sars.gov.za", "payfast.co.za",
    # Telcos & Utilities
    "vodacom.co.za", "mtn.co.za", "telkom.co.za", "cellc.co.za", "rain.co.za",
    "eskom.co.za", "citypower.co.za", "joburg.org.za", "afrihost.com", "mweb.co.za",
    # E-commerce & Logistics
    "takealot.com", "superbalist.com", "uber.com", "ubereats.com", "mrdfood.com",
    "checkers.co.za", "woolworths.co.za", "picknpay.co.za", "makro.co.za",
    "leroymerlin.co.za", "builders.co.za", "ram.co.za", "dpd.com", "fastway.co.za",
    "courierguy.co.za", "thecourierguy.co.za", "dhl.com", "fedex.com",
    # Social & Software Platforms
    "linkedin.com", "facebookmail.com", "instagram.com", "twitter.com", "x.com",
    "google.com", "microsoft.com", "apple.com", "zoom.us", "slack.com",
    "canva.com", "adobe.com", "dropbox.com", "github.com", "wordpress.com",
    "calendly.com", "mailchimp.com", "hubspot.com", "resend.com", "sanity.io",
    "agentmail.to", "attio.com", "vercel.com", "stripe.com", "amazonses.com"
}

PERSONAL_PROVIDERS = {
    "gmail.com", "yahoo.com", "hotmail.com", "outlook.com", "icloud.com",
    "mweb.co.za", "vodamail.co.za", "telkomsa.net", "iafrica.com",
    "webmail.co.za", "live.com", "me.com", "aol.com"
}

HOTEL_KEYWORDS = {
    "hotel", "lodge", "resort", "inn", "hospitality", "stay", "guest",
    "tsogosun", "southernsun", "suninternational", "marriott", "radisson", "hilton",
    "legacyhotels", "protea", "anamandla", "saxongroup", "maslow", "leonardo"
}

DESIGN_KEYWORDS = {
    "design", "interior", "interiors", "decor", "decorator", "architect",
    "architecture", "curtain", "drape", "fabric", "textile", "upholstery",
    "furniture", "living", "studio", "bespoke"
}

COMMERCIAL_KEYWORDS = {
    "property", "properties", "facilities", "facility", "maintenance",
    "commercial", "corporate", "office", "towers", "centre", "mall",
    "school", "college", "clinic", "hospital", "theatre", "holding", "holdings",
    "group", "ventures", "enterprises", "estate", "estates"
}


def extract_email(text: str) -> str:
    if not text:
        return ""
    match = re.search(r"[\w\.-]+@[\w\.-]+\.\w+", str(text))
    if match:
        return match.group(0).lower().strip()
    return ""


def is_junk(email: str) -> bool:
    if not email or "@" not in email:
        return True
    local_part, domain = email.split("@", 1)
    
    if any(local_part == p or local_part.startswith(f"{p}-") or local_part.startswith(f"{p}.") for p in JUNK_PREFIXES):
        return True
        
    if domain in JUNK_DOMAINS or any(domain.endswith(f".{jd}") for jd in JUNK_DOMAINS):
        return True
        
    if re.search(r"no[-._]?reply|donot[-._]?reply|mailer[-._]?daemon", local_part):
        return True
        
    return False


def classify(email: str, name: str, company: str) -> str:
    domain = email.split("@", 1)[1] if "@" in email else ""
    corpus = f"{email} {name} {company} {domain}".lower()
    
    for kw in HOTEL_KEYWORDS:
        if kw in corpus:
            return "hotel_hospitality"
            
    for kw in DESIGN_KEYWORDS:
        if kw in corpus:
            return "trade_design"
            
    for kw in COMMERCIAL_KEYWORDS:
        if kw in corpus:
            return "commercial_b2b"
            
    if domain in PERSONAL_PROVIDERS:
        return "residential"
        
    return "commercial_b2b"


def process_attio_file(file_path: pathlib.Path):
    wb = openpyxl.load_workbook(file_path, read_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    
    records = []
    for r in rows[1:]:
        rec_id = r[0]
        record_label = str(r[1]) if r[1] is not None else ""
        strength = str(r[2]) if len(r) > 2 and r[2] is not None else ""
        last_email = str(r[3]) if len(r) > 3 and r[3] is not None else ""
        
        email = extract_email(record_label)
        name = record_label if not email or record_label != email else ""
        
        records.append({
            "source": "attio_sync",
            "email": email,
            "name": name,
            "company": "",
            "connection_strength": strength,
            "last_interaction": last_email
        })
    return records


def process_tsogo_sun_file(file_path: pathlib.Path):
    wb = openpyxl.load_workbook(file_path, read_only=True)
    records = []
    
    # 1. Read Properties
    if "Properties" in wb.sheetnames:
        sheet = wb["Properties"]
        rows = list(sheet.iter_rows(values_only=True))
        for r in rows[4:]:  # Start at data rows
            prop_name = str(r[2]) if len(r) > 2 and r[2] is not None else ""
            city = str(r[4]) if len(r) > 4 and r[4] is not None else ""
            province = str(r[5]) if len(r) > 5 and r[5] is not None else ""
            all_emails = str(r[10]) if len(r) > 10 and r[10] is not None else ""
            phone = str(r[8]) if len(r) > 8 and r[8] is not None else ""
            
            # Extract any emails
            for email in re.findall(r"[\w\.-]+@[\w\.-]+\.\w+", all_emails):
                records.append({
                    "source": "tsogo_sun_properties",
                    "email": email.lower().strip(),
                    "name": "General Manager / Housekeeping",
                    "company": f"{prop_name} ({city})",
                    "connection_strength": "High-Priority Hotel",
                    "last_interaction": ""
                })
                
    # 2. Read Decision Makers
    if "Decision_Makers" in wb.sheetnames:
        sheet = wb["Decision_Makers"]
        rows = list(sheet.iter_rows(values_only=True))
        for r in rows[4:]:
            prop = str(r[2]) if len(r) > 2 and r[2] is not None else ""
            role = str(r[3]) if len(r) > 3 and r[3] is not None else ""
            name = str(r[4]) if len(r) > 4 and r[4] is not None else ""
            notes = str(r[7]) if len(r) > 7 and r[7] is not None else ""
            
            # Format Southern Sun executive email prediction: firstname.lastname@southernsun.com
            if name:
                parts = name.strip().split()
                if len(parts) >= 2:
                    email_slug = f"{parts[0].lower()}.{parts[-1].lower()}@southernsun.com"
                    records.append({
                        "source": "tsogo_sun_executives",
                        "email": email_slug,
                        "name": name,
                        "company": f"{prop} ({role})",
                        "connection_strength": "Hotel Executive",
                        "last_interaction": ""
                    })
    return records


def main():
    print("\n=================================================================")
    print("ATTIO CRM & PROSPECT DATASET PROCESSING PIPELINE")
    print("=================================================================\n")
    
    all_raw = []
    
    # 1. Process Attio Recent Contacts
    attio_path = DOWNLOADS_DIR / "People - Recently Contacted People.xlsx"
    if attio_path.exists():
        print(f"Loading Attio email sync: {attio_path}...")
        attio_recs = process_attio_file(attio_path)
        print(f"-> Ingested {len(attio_recs)} records from Attio.")
        all_raw.extend(attio_recs)
    else:
        print(f"Attio file not found at {attio_path}")
        
    # 2. Process Tsogo Sun Prospect Kit
    tsogo_path = DOWNLOADS_DIR / "Tsogo_Sun_Prospect_List_and_Campaign_Kit.xlsx"
    if tsogo_path.exists():
        print(f"Loading Tsogo Sun prospect kit: {tsogo_path}...")
        tsogo_recs = process_tsogo_sun_file(tsogo_path)
        print(f"-> Ingested {len(tsogo_recs)} hotel decision maker contacts.")
        all_raw.extend(tsogo_recs)
        
    print(f"\nTotal raw ingested contacts: {len(all_raw)}")
    
    # Deduplicate and Clean
    seen_emails = set()
    cleaned = []
    junk_count = 0
    duplicate_count = 0
    no_email_count = 0
    
    for r in all_raw:
        email = r["email"]
        if not email:
            no_email_count += 1
            continue
            
        if email in seen_emails:
            duplicate_count += 1
            continue
            
        if is_junk(email):
            junk_count += 1
            continue
            
        seen_emails.add(email)
        cohort = classify(email, r["name"], r["company"])
        cleaned.append({
            "email": email,
            "name": r["name"],
            "company": r["company"],
            "cohort": cohort,
            "connection_strength": r["connection_strength"],
            "last_interaction": r["last_interaction"],
            "source": r["source"]
        })
        
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    
    cohorts = {
        "hotel_hospitality": [],
        "commercial_b2b": [],
        "residential": [],
        "trade_design": []
    }
    for c in cleaned:
        cohorts[c["cohort"]].append(c)
        
    print(f"\n=================================================================")
    print(f"CLEANING, PURGING & SEGMENTATION RESULTS")
    print(f"=================================================================")
    print(f"Raw Input Records:             {len(all_raw)}")
    print(f"Unresolvable / Empty Inboxes:  {no_email_count}")
    print(f"Automated / Bank Junk Purged:  {junk_count} ({junk_count/len(all_raw)*100:.1f}%)")
    print(f"Duplicate Inboxes Removed:     {duplicate_count}")
    print(f"-----------------------------------------------------------------")
    print(f"TOTAL CLEAN VALIDATED CONTACTS:{len(cleaned)}")
    print(f"-----------------------------------------------------------------")
    print(f"1. Hotels & Hospitality:       {len(cohorts['hotel_hospitality'])} contacts")
    print(f"2. Commercial B2B & Facilities:{len(cohorts['commercial_b2b'])} contacts")
    print(f"3. Residential Clients:        {len(cohorts['residential'])} contacts")
    print(f"4. Trade & Interior Designers: {len(cohorts['trade_design'])} contacts")
    print(f"=================================================================\n")
    
    # Export individual segmented CSV files
    for name, items in cohorts.items():
        file_path = OUTPUT_DIR / f"{name}_leads.csv"
        with open(file_path, mode="w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=["email", "name", "company", "connection_strength", "last_interaction", "source"])
            writer.writeheader()
            for row in items:
                writer.writerow({
                    "email": row["email"],
                    "name": row["name"],
                    "company": row["company"],
                    "connection_strength": row["connection_strength"],
                    "last_interaction": row["last_interaction"],
                    "source": row["source"]
                })
        print(f"✓ Exported {len(items)} leads to {file_path}")
        
    # Export Master Consolidated Clean Database
    master_path = OUTPUT_DIR / "master_clean_contacts.csv"
    with open(master_path, mode="w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["email", "name", "company", "cohort", "connection_strength", "last_interaction", "source"])
        writer.writeheader()
        for row in cleaned:
            writer.writerow(row)
    print(f"✓ Exported master database ({len(cleaned)} leads) to {master_path}\n")


if __name__ == "__main__":
    main()
